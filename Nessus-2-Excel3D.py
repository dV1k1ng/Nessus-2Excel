#!/usr/bin/env python
# Author: Erik A - dV1k1ng
# Description: Create a 2-sheet-Excel-from-nessus-output;
# Sheet 1 'overview' contains a table issues vs IP addresses, cells are protocol/ports
# Sheet 2 'details' contains the issues with details
# Clicking an issue on sheet 1 jumps to the details of that issue on sheet 2


import argparse
import sys
from operator import itemgetter, attrgetter, methodcaller

import xml.etree.ElementTree as etree
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
import shutil
import os
import json

def ExtractIp(lst):
    return [item['ip'] for item in lst]

# table with names and color codes of each severity level
severity = {
    0: {'name':"Informational", 'color': "67A64D"},
    1: {'name':"Low",           'color': "FFFF00"},
    2: {'name':"Medium",        'color': "F8C346"},
    3: {'name':"High",          'color': "FF0000"},
    4: {'name':"Critical",      'color': "8B0000"}
}

# parse arguments
parser = argparse.ArgumentParser(
    description='This script will create a 3D Excel sheet from a Nessus scan')
parser.add_argument('--nessus',
                    required=True,
                    help="Nessus file")
parser.add_argument('--output',
                    required=True,
                    help="Output filename (without extension)")

args = parser.parse_args()

filename_nessus = args.nessus
filename_output = args.output

filename_output_excel = filename_output + '.xlsx'

if not filename_nessus.endswith('.nessus'):
    print ("Nessus file does not end with .nessus extension")
    parser.print_help()
    sys.exit(1)


# create empty plugin lookup table and empty findings
plugins = {}
hosts = {}

# load nessus file
tree = etree.parse(filename_nessus)
for host in tree.findall('.//ReportHost'):
    ip = host.find('HostProperties/tag[@name="host-ip"]').text
    hostname = host.attrib['name']

    for item in host.findall('ReportItem'):
        severityid = int(item.attrib['severity'])
        pluginid = str(item.attrib['pluginID'])
        port = item.attrib['port']
        protocol = str(item.attrib['protocol'])
        pluginName = str(item.attrib['pluginName'])
        for desc in item.findall('description'):
            description = desc.text
        for sol in item.findall('solution'):
            solution = sol.text
        for evi in item.findall('plugin_output'):
            evidence = evi.text

#       print into plugins arry
        if not pluginid in plugins:
            # new pluginid, initialize row in array
            plugins[pluginid] = {'pluginid':pluginid,'instances':0,'severityid':severityid,'title':pluginName,'description':description,'solution':solution,'evidence':evidence,'hosts':[]}
        plugins[pluginid]['instances'] += 1
        plugins[pluginid]['hosts'].append({'hostname':hostname,'ip':ip,'port':port,'protocol':protocol})

#       print into hosts array
        if not ip in hosts:
            # new host, initialize row in array
            hosts[ip] = {'ip':ip,'hostname':hostname,'issues':[]}
        # 2020-02-14 (EA): more details into hosts issues array
        hosts[ip]['issues'].append({'port':port,'protocol':protocol,'pluginName':pluginName,'severityid':severityid,'description':description,'solution':solution,'evidence':evidence})

# convert to array and sort descending
plugins = sorted(iter(plugins.values()), key=lambda k: k['instances'], reverse=True)
hosts = sorted(iter(hosts.values()), key=lambda k: k['ip'], reverse=False)

# create a new Excel sheet
workbook = Workbook()
sheet1=workbook.active
sheet1.title="Overview"
sheet2=workbook.create_sheet("Details")

header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="left", vertical="top")
MaxWidthB = 1
MaxWidthIP = 15

# write ip's into first row
Cols = ExtractIp(hosts)
#print(Cols)
NrOfHosts = len(Cols)
sheet1.append(Cols)
# and insert two empty column before, for issue severity and title
ColOffset = 2
sheet1.insert_cols(idx=1)
sheet1.insert_cols(idx=1)

#sheet2
Cols2 = ['Name', 'Description', 'Solution', 'Explanation','Severity']
NrOfCols = len(Cols2)
sheet2.append(Cols2)

for severityid in [4,3,2,1]:
    for plugin in plugins:
        if plugin['severityid'] == severityid and plugin['instances'] > 0:
            # new relevant issue
            ExcelRow = [''] * (NrOfHosts+ColOffset)
            ExcelRow[0] = severity[plugin['severityid']]['name']
            ExcelRow[1] = plugin['title']
            # remember longest entry in this column
            if len(ExcelRow[1]) > MaxWidthB:
                MaxWidthB = len(ExcelRow[1])

            # write details into sheet2
            ExcelRow2 = [''] * 5
            ExcelRow2[4] = severity[plugin['severityid']]['name']
            ExcelRow2[0] = plugin['title']
            ExcelRow2[1] = plugin['description']
            ExcelRow2[2] = plugin['solution']
            ExcelRow2[3] = plugin['evidence']
            sheet2.append(ExcelRow2)

            # enumerate IP's for this issue
            for host in plugin['hosts']:
                # Look for IP in Initial row (i.e. column headers), and get index
                index = Cols.index(host['ip'])
                if (ExcelRow[index+ColOffset]) != '':
                    ExcelRow[index+ColOffset] += ', \r\n'
                ExcelRow[index+ColOffset] += str(host['port']) + '/' + str(host['protocol'])

            sheet1.append(ExcelRow)


print ("Writing output to files: ", filename_output)

# Formatting of Excel worksheet
sheet1.column_dimensions['B'].width = MaxWidthB
for i, column_cells in enumerate(sheet1.columns, start=3):
    col = get_column_letter(i)
    sheet1.column_dimensions[col].width = MaxWidthIP

header_row = sheet1[1]
for cell in header_row:
    cell.style = header
header_row = sheet1["B"]
for cell in header_row:
    cell.style = header

critical_text = PatternFill(bgColor=severity[4]['color'])
high_text = PatternFill(bgColor=severity[3]['color'])
medium_text = PatternFill(bgColor=severity[2]['color'])
low_text = PatternFill(bgColor=severity[1]['color'])
alignment_text = Alignment(vertical='top', horizontal='left')

critical_diff_style = DifferentialStyle(fill=critical_text, alignment=alignment_text)
high_diff_style = DifferentialStyle(fill=high_text, alignment=alignment_text)
medium_diff_style = DifferentialStyle(fill=medium_text, alignment=alignment_text)
low_diff_style = DifferentialStyle(fill=low_text, alignment=alignment_text)

rule = Rule(type="expression", dxf=critical_diff_style)
rule.formula = ['NOT(ISERROR(SEARCH("Critical",$A1)))']
sheet1.conditional_formatting.add("A1:A1000", rule)

rule = Rule(type="expression", dxf=high_diff_style)
rule.formula = ['NOT(ISERROR(SEARCH("High",$A1)))']
sheet1.conditional_formatting.add("A1:A1000", rule)

rule = Rule(type="expression", dxf=medium_diff_style)
rule.formula = ['NOT(ISERROR(SEARCH("Medium",$A1)))']
sheet1.conditional_formatting.add("A1:A1000", rule)

rule = Rule(type="expression", dxf=low_diff_style)
rule.formula = ['NOT(ISERROR(SEARCH("Low",$A1)))']
sheet1.conditional_formatting.add("A1:A1000", rule)

# Below cell auto-wrap
for rows in sheet1.iter_rows(min_row=2, min_col=3):
    for cell in rows:
        cell.alignment = Alignment(wrapText=True)

sheet1.freeze_panes = "C2"

header_row = sheet2[1]
for cell in header_row:
    cell.style = header

sheet2.column_dimensions['A'].width = 20
sheet2.column_dimensions['B'].width = 80
sheet2.column_dimensions['C'].width = 40
sheet2.column_dimensions['D'].width = 80
sheet2.column_dimensions['E'].width = 10

for rows in sheet2.iter_rows(min_row=2, min_col=1):
    for cell in rows:
        cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')

sheet2.freeze_panes = "A2"

# make hyperlinks from sheet1 to sheet2
row = 1
for rows in sheet1.iter_rows(min_row=2, min_col=2, max_col=2):
    row = row + 1
    for cell in rows:
        link = '#Details!A' + str(row)
        cell.hyperlink = link

# Save worksheet
workbook.save(filename=filename_output_excel)

print ("Done.")
